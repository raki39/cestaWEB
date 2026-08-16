"""
Gerador do modelo `.xlsx` que o cliente baixa.

Por que gerado, e não um arquivo commitado
------------------------------------------
O modelo é construído a partir do MESMO dicionário `COLUNAS` que o parser usa.
Assim ele não pode divergir: não existe o estado "o modelo pede uma coluna que o
parser não conhece" — que é o jeito clássico de o cliente receber "faltou a
coluna 'preco'" numa planilha que ele baixou de nós.

A coluna de EAN sai formatada como **texto** (`numFmt = '@'`). É a única defesa
real contra o Excel transformar `7891107101012` em `7,89111E+12`: quando isso
acontece no arquivo, os dígitos do fim já se perderam e o parser recusa o valor
(ver `planilha._limpa_ean`).
"""
from __future__ import annotations

import io
from typing import Dict, List

from .planilha import COLUNAS, OBRIGATORIAS

NOME_ARQUIVO = "modelo-cesta.xlsx"

# Linhas de exemplo, para o cliente ver o formato esperado em vez de deduzir.
# Por campo (dict, não posição): a ordem das colunas na planilha é decidida em
# `gerar_modelo`, e casar por posição aqui já tinha me feito escrever o preço na
# coluna do EAN.
EXEMPLO: List[Dict[str, object]] = [
    {"descricao": "OLEO DE SOJA SOYA 900ML", "ean": "7891107101012",
     "preco": 7.49, "marca": "SOYA", "categoria": "MERCEARIA"},
    {"descricao": "CAFE TORRADO E MOIDO 500G", "ean": "7896005800011",
     "preco": 15.90, "marca": "MELITTA", "categoria": "MERCEARIA"},
    {"descricao": "ARROZ BRANCO TIPO 1 5KG", "preco": 24.90},
]


def gerar_modelo() -> bytes:
    """Bytes do `.xlsx` modelo, montado a partir de `COLUNAS`."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    campos = list(COLUNAS.keys())          # ordem de declaração: descricao, preco, ean, ...
    # Ordem de exibição pensada para quem preenche: o obrigatório primeiro.
    ordem = [c for c in ("descricao", "preco", "ean", "marca", "categoria") if c in campos]

    wb = openpyxl.Workbook()
    aba = wb.active
    aba.title = "cesta"

    cabecalho = Font(bold=True, color="FFFFFF")
    fundo = PatternFill("solid", fgColor="2D5D4B")
    fundo_opcional = PatternFill("solid", fgColor="7A8C84")

    for i, campo in enumerate(ordem, start=1):
        c = aba.cell(row=1, column=i, value=campo)
        c.font = cabecalho
        c.fill = fundo if campo in OBRIGATORIAS else fundo_opcional
        c.alignment = Alignment(horizontal="left")

    for linha_idx, linha in enumerate(EXEMPLO, start=2):
        for i, campo in enumerate(ordem, start=1):
            aba.cell(row=linha_idx, column=i, value=linha.get(campo))

    # Formatação por coluna. O texto no EAN é o ponto todo deste arquivo.
    for i, campo in enumerate(ordem, start=1):
        letra = get_column_letter(i)
        largura = {"descricao": 46, "preco": 12, "ean": 18, "marca": 20, "categoria": 20}
        aba.column_dimensions[letra].width = largura.get(campo, 18)
        if campo == "ean":
            # Vale para a coluna inteira, não só para as linhas de exemplo: é o que
            # faz o Excel tratar o que o cliente digitar como texto.
            for linha_idx in range(2, 2000):
                aba.cell(row=linha_idx, column=i).number_format = "@"
        elif campo == "preco":
            for linha_idx in range(2, 2000):
                aba.cell(row=linha_idx, column=i).number_format = "#,##0.00"

    aba.freeze_panes = "A2"

    # Aba de instruções: quem abre a planilha lê aqui o que é obrigatório e por que
    # o EAN é texto, sem precisar do e-mail que veio junto.
    ajuda = wb.create_sheet("como preencher")
    ajuda.column_dimensions["A"].width = 100
    linhas_ajuda = [
        "COMO PREENCHER A CESTA",
        "",
        "1. Preencha uma linha por produto do seu tabloide, na aba 'cesta'.",
        "2. Colunas obrigatórias: " + ", ".join(OBRIGATORIAS) + ".",
        "3. As demais colunas são opcionais — pode deixar em branco.",
        "4. A ORDEM das linhas é a ordem do seu tabloide. É ela que define o TOP N.",
        "",
        "SOBRE O CÓDIGO DE BARRAS (EAN)",
        "A coluna já vem formatada como TEXTO. Não mude isso: se o Excel tratar o",
        "código como número, ele vira 7,89111E+12 e os últimos dígitos são perdidos",
        "dentro do próprio arquivo. Nesse caso o código é descartado no envio (com",
        "aviso na tela) — o produto entra na cesta, só sem o EAN.",
        "",
        "SOBRE O PREÇO",
        "Use o preço anunciado no seu tabloide. Aceita 7,49 ou 7.49.",
        "Preço zerado, negativo ou em branco derruba a linha (avisamos qual).",
        "",
        "NÃO RENOMEIE AS COLUNAS DA PRIMEIRA LINHA.",
        "Se renomear, o envio não vai encontrar o cabeçalho.",
    ]
    for i, texto in enumerate(linhas_ajuda, start=1):
        c = ajuda.cell(row=i, column=1, value=texto)
        if i == 1 or texto.isupper() and texto:
            c.font = openpyxl.styles.Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()
