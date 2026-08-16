"""
Leitura da planilha de cesta — o único lugar do projeto que abre `.xlsx`.

Este módulo é a razão de o `cestaWEB` existir separado do OCTO: abrir e
normalizar planilha é trabalho de CPU, e ele não pode acontecer no container que
atende o chat. Aqui ele acontece longe.

Contrato de saída
-----------------
`ler_planilha()` devolve `ResultadoLeitura`, com:

    itens   — lista de dicts JÁ no formato dos itens do `POST /cestas`
              (ver a whitelist em CONTRATO.md; nenhum campo extra sai daqui)
    avisos  — o que foi descartado ou corrigido, em português, para a tela
    resumo  — contagens para a tela de resultado

O que NÃO acontece aqui
-----------------------
Casamento com a base de mercado. Este módulo não sabe o que é `NM_PRODUTO`, não
conhece ClickHouse e não decide se um produto existe. A descrição sai como o
cliente escreveu (normalizada) e quem resolve é o OCTO.

Princípio geral: **nada é descartado em silêncio**. Toda linha que não vira item
gera um aviso com o número da linha do Excel — quem subiu a planilha precisa
saber o que ficou de fora, senão o relatório vem certo respondendo outra
pergunta.
"""
from __future__ import annotations

import logging
import re
import unicodedata
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

# ── Limites ─────────────────────────────────────────────────────────────────────
# Itens por cesta. Protege o parser e a tela de cestas do OCTO. NÃO é
# o mesmo teto que limita a SIMULAÇÃO (esse é o TOP N, decidido dentro do OCTO, e
# é bem menor: a execução dispara consultas por produto).
MAX_ITENS = 5_000
# Linhas que o parser percorre antes de desistir. Maior que MAX_ITENS de propósito:
# planilha real tem linha em branco, subtotal e rodapé no meio.
MAX_LINHAS_VARRIDAS = 20_000
# Quantas linhas do topo podem ser título/logo antes do cabeçalho de verdade.
MAX_LINHAS_ATE_CABECALHO = 15

TAMANHOS_EAN_VALIDOS = (8, 12, 13, 14)

LIMITE_DESCRICAO = 500
LIMITE_TEXTO_CURTO = 200


# ── Colunas aceitas ─────────────────────────────────────────────────────────────
# Cabeçalho é casado sem acento, sem caixa e sem espaço duplo. O primeiro alias
# de cada tupla é o nome que está no modelo .xlsx distribuído ao cliente.
COLUNAS: Dict[str, Tuple[str, ...]] = {
    "descricao": ("descricao", "descricao do produto", "produto", "item",
                  "descricao item", "nome do produto"),
    "preco": ("preco", "preco promocional", "preco oferta", "valor", "preco venda"),
    "ean": ("ean", "codigo de barras", "cod barras", "gtin", "codigo"),
    "marca": ("marca", "fabricante"),
    "categoria": ("categoria", "depto", "departamento", "secao", "setor"),
}
OBRIGATORIAS = ("descricao", "preco")

# Alias → campo canônico, resolvido uma vez.
_ALIAS_PARA_CAMPO: Dict[str, str] = {
    alias: campo for campo, aliases in COLUNAS.items() for alias in aliases
}


@dataclass
class ResultadoLeitura:
    itens: List[Dict[str, Any]] = field(default_factory=list)
    avisos: List[str] = field(default_factory=list)
    resumo: Dict[str, Any] = field(default_factory=dict)

    @property
    def ok(self) -> bool:
        return bool(self.itens)


class PlanilhaInvalida(Exception):
    """A planilha não pode ser lida — mensagem já pronta para a tela.

    Separada dos avisos de propósito: aviso é "essa linha ficou de fora", erro é
    "não há o que ler". Confundir os dois faria uma planilha vazia parecer aceita.
    """


# ── Normalização de texto ───────────────────────────────────────────────────────
def _sem_acento(s: Any) -> str:
    t = unicodedata.normalize("NFKD", str(s or ""))
    return t.encode("ascii", "ignore").decode()


def _norm_cabecalho(s: Any) -> str:
    """Cabeçalho comparável: sem acento, minúsculo, espaços colapsados, sem
    pontuação de borda ('Preço (R$):' → 'preco r$')."""
    t = _sem_acento(s).lower().replace("_", " ").replace("-", " ")
    t = re.sub(r"[.:;,]", " ", t)
    return re.sub(r"\s+", " ", t).strip()


def _limpa_texto(v: Any, limite: int) -> Optional[str]:
    """Texto de célula → string enxuta, ou None se não sobrar nada.

    Mantém acento (é dado do cliente, não chave de comparação) e colapsa espaço,
    que em planilha vem sujo de copiar-e-colar.
    """
    if v is None:
        return None
    t = re.sub(r"\s+", " ", str(v)).strip()
    if not t:
        return None
    return t[:limite]


# ── EAN ─────────────────────────────────────────────────────────────────────────
def _limpa_ean(v: Any) -> Tuple[Optional[str], Optional[str]]:
    """Valor da célula → (ean, motivo_do_descarte).

    O caso que ESTE código existe para resolver: o Excel guarda `7891107101012`
    como float e a célula chega aqui como `7.891107101012e+12`. `str()` nisso
    devolve notação científica, e gravar isso no banco é gravar lixo. Então:

      - numérico vira inteiro (sem expoente, sem separador de milhar);
      - só dígitos sobrevivem (tira espaço, ponto, hífen);
      - o tamanho tem de ser de EAN de verdade — senão devolve None COM motivo.

    Zero à esquerda que o Excel comeu não tem volta aqui (a informação já se
    perdeu no arquivo); é por isso que a coluna do modelo vem formatada como
    texto. O que este código garante é que o dado errado não passa calado.
    """
    if v is None:
        return None, None

    if isinstance(v, bool):
        return None, "não é um código"

    if isinstance(v, (int, float, Decimal)):
        try:
            inteiro = int(Decimal(str(v)))
        except (InvalidOperation, ValueError, OverflowError):
            return None, "número que não pôde ser lido"
        bruto = str(inteiro)
    else:
        bruto = str(v).strip()
        # "7,89111E+12" — a célula já foi salva EM notação científica, então os
        # dígitos do fim não existem mais no arquivo. Reconstruir daria
        # '7891110000000': treze dígitos, formato válido e ERRADO (o EAN real era
        # 7891107101012). Dá para detectar comparando os dígitos significativos da
        # mantissa com o tamanho do resultado — se faltam, o Excel comeu a cauda e
        # nós estaríamos fabricando zeros. Aí é recusa, não conserto.
        if re.fullmatch(r"[\d.,]+[eE][+-]?\d+", bruto):
            try:
                inteiro = int(Decimal(bruto.replace(",", ".")))
            except (InvalidOperation, ValueError, OverflowError):
                return None, "notação científica que não pôde ser lida"
            # Dígitos que REALMENTE estão escritos na mantissa contra o tamanho do
            # inteiro expandido. Se a mantissa não dá conta de todas as casas, as
            # que faltam foram preenchidas com zero pelo expoente — não são dado.
            # (Nada de `rstrip('0')` aqui: era o que fazia a perda desaparecer na
            #  própria comparação, e um EAN pode legitimamente terminar em zero.)
            significativos = re.sub(r"\D", "", bruto.split("e")[0].split("E")[0])
            if len(significativos) < len(str(inteiro)):
                return None, (
                    "veio em notação científica e os dígitos do fim foram perdidos pelo Excel "
                    "— formate a coluna como TEXTO e envie de novo"
                )
            bruto = str(inteiro)

    digitos = re.sub(r"\D", "", bruto)
    if not digitos:
        return None, "sem dígitos"
    digitos = digitos.lstrip("0") or "0"
    if len(digitos) not in TAMANHOS_EAN_VALIDOS:
        return None, f"{len(digitos)} dígitos (esperado 8, 12, 13 ou 14)"
    return digitos, None


# ── Preço ───────────────────────────────────────────────────────────────────────
def _limpa_preco(v: Any) -> Tuple[Optional[float], Optional[str]]:
    """Valor da célula → (preço, motivo_do_descarte).

    Aceita o que aparece de verdade numa planilha brasileira: `7,49`, `R$ 7,49`,
    `1.234,56`, e o float que o Excel já entrega pronto.
    """
    if v is None:
        return None, "vazio"

    if isinstance(v, bool):
        return None, "não é um valor"

    if isinstance(v, (int, float, Decimal)):
        valor = float(v)
    else:
        t = str(v).strip()
        if not t:
            return None, "vazio"
        t = re.sub(r"(?i)\s*r\$\s*", "", t)
        t = re.sub(r"[^\d.,-]", "", t)
        if not t:
            return None, "sem número"
        # Decide quem é separador decimal: o ÚLTIMO separador que aparece manda.
        # '1.234,56' → vírgula decimal;  '1,234.56' → ponto decimal;  '7,49' → vírgula.
        ultima_virgula, ultimo_ponto = t.rfind(","), t.rfind(".")
        if ultima_virgula > ultimo_ponto:
            t = t.replace(".", "").replace(",", ".")
        else:
            t = t.replace(",", "")
        try:
            valor = float(t)
        except ValueError:
            return None, f"valor não numérico ({str(v).strip()[:20]})"

    if valor != valor or valor in (float("inf"), float("-inf")):  # NaN / infinito
        return None, "valor inválido"
    if valor <= 0:
        return None, "preço menor ou igual a zero"
    return round(valor, 2), None


# ── Cabeçalho ───────────────────────────────────────────────────────────────────
def _acha_cabecalho(linhas: List[Tuple[Any, ...]]) -> Tuple[int, Dict[str, int], List[str]]:
    """Acha a linha de cabeçalho e o mapa campo → índice de coluna.

    Varre as primeiras linhas em vez de assumir a linha 1: planilha de cliente
    costuma ter título, logo ou linha em branco no topo. Vence a primeira linha
    que traga TODAS as colunas obrigatórias.

    Devolve (índice_da_linha, {campo: coluna}, nomes_ignorados).
    """
    melhor_erro = "nenhuma linha com as colunas obrigatórias"

    for idx, linha in enumerate(linhas[:MAX_LINHAS_ATE_CABECALHO]):
        mapa: Dict[str, int] = {}
        ignorados: List[str] = []
        for col, valor in enumerate(linha):
            nome = _norm_cabecalho(valor)
            if not nome:
                continue
            campo = _ALIAS_PARA_CAMPO.get(nome)
            if campo is None:
                # Tolerância: 'preco (r$)' contém 'preco'. Só para os alias longos,
                # para 'codigo' não roubar a coluna de 'codigo interno'.
                for alias, cand in _ALIAS_PARA_CAMPO.items():
                    if len(alias) >= 5 and nome.startswith(alias):
                        campo = cand
                        break
            if campo is None:
                ignorados.append(str(valor).strip())
            elif campo not in mapa:            # primeira coluna vence a duplicata
                mapa[campo] = col

        faltando = [c for c in OBRIGATORIAS if c not in mapa]
        if not faltando:
            return idx, mapa, ignorados
        if idx == 0 and any(_norm_cabecalho(v) for v in linha):
            melhor_erro = "faltou a coluna " + " e ".join(f"'{c}'" for c in faltando)

    raise PlanilhaInvalida(
        f"Não encontrei o cabeçalho da planilha: {melhor_erro}. "
        f"Baixe o modelo e mantenha a primeira linha como está."
    )


# ── Leitura ─────────────────────────────────────────────────────────────────────
def ler_planilha(conteudo: bytes, nome_arquivo: str = "planilha.xlsx") -> ResultadoLeitura:
    """Bytes do `.xlsx` → itens prontos para o `POST /cestas`.

    Levanta `PlanilhaInvalida` quando não há o que ler (arquivo corrompido, aba
    vazia, cabeçalho ausente, nenhuma linha aproveitável). Linha ruim isolada não
    interrompe: entra em `avisos` e a leitura continua.
    """
    try:
        import openpyxl
    except ImportError as e:  # pragma: no cover - dependência declarada
        raise PlanilhaInvalida("Leitor de planilha não disponível no servidor.") from e

    import io

    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    except Exception as e:
        logger.warning("[PLANILHA] falha ao abrir %s: %s", nome_arquivo, e)
        raise PlanilhaInvalida(
            "Não consegui abrir o arquivo. Envie um .xlsx salvo pelo Excel "
            "(arquivo .xls antigo, .csv renomeado ou arquivo corrompido não servem)."
        ) from e

    try:
        aba = wb[wb.sheetnames[0]] if wb.sheetnames else None
        if aba is None:
            raise PlanilhaInvalida("A planilha não tem nenhuma aba.")

        linhas: List[Tuple[Any, ...]] = []
        for i, linha in enumerate(aba.iter_rows(values_only=True)):
            if i >= MAX_LINHAS_VARRIDAS:
                break
            linhas.append(linha)
    finally:
        wb.close()

    if not linhas:
        raise PlanilhaInvalida("A primeira aba da planilha está vazia.")

    idx_cab, mapa, ignorados = _acha_cabecalho(linhas)
    logger.info("[PLANILHA] %s | cabeçalho na linha %d | colunas=%s",
                nome_arquivo, idx_cab + 1, sorted(mapa))

    res = ResultadoLeitura()
    if ignorados:
        res.avisos.append(
            "Colunas ignoradas (não fazem parte do modelo): "
            + ", ".join(sorted(set(ignorados))[:8])
        )
    for opcional in ("ean", "marca", "categoria"):
        if opcional not in mapa:
            res.avisos.append(f"A planilha não tem a coluna '{opcional}'. Segui sem ela.")

    def celula(linha: Tuple[Any, ...], campo: str) -> Any:
        col = mapa.get(campo)
        if col is None or col >= len(linha):
            return None
        return linha[col]

    descartadas = 0
    vistas: Dict[Tuple[str, Optional[str]], int] = {}   # (descrição, ean) → linha do Excel
    duplicadas = 0
    ean_descartados = 0
    truncou_no_teto = False

    for offset, linha in enumerate(linhas[idx_cab + 1:]):
        n_excel = idx_cab + 2 + offset      # 1-based, como o Excel mostra

        if not any(v is not None and str(v).strip() for v in linha):
            continue                        # linha em branco: nem conta

        descricao = _limpa_texto(celula(linha, "descricao"), LIMITE_DESCRICAO)
        preco, motivo_preco = _limpa_preco(celula(linha, "preco"))

        if descricao is None:
            descartadas += 1
            if descartadas <= 20:
                res.avisos.append(f"Linha {n_excel}: sem descrição — ficou de fora.")
            continue
        if preco is None:
            descartadas += 1
            if descartadas <= 20:
                res.avisos.append(f"Linha {n_excel} ({descricao[:40]}): {motivo_preco} — ficou de fora.")
            continue

        ean, motivo_ean = _limpa_ean(celula(linha, "ean"))
        if motivo_ean:
            ean_descartados += 1
            if ean_descartados <= 10:
                res.avisos.append(
                    f"Linha {n_excel} ({descricao[:40]}): código de barras descartado — {motivo_ean}. "
                    "O produto entrou na cesta; só o EAN ficou vazio."
                )

        chave = (_sem_acento(descricao).upper(), ean)
        if chave in vistas:
            duplicadas += 1
            if duplicadas <= 10:
                res.avisos.append(
                    f"Linha {n_excel} ({descricao[:40]}): repetida da linha {vistas[chave]} — mantive a primeira."
                )
            continue
        vistas[chave] = n_excel

        if len(res.itens) >= MAX_ITENS:
            truncou_no_teto = True
            break

        res.itens.append({
            "ordem": len(res.itens) + 1,
            "descricao_cliente": descricao,
            "ean": ean,
            "preco": preco,
            "marca": _limpa_texto(celula(linha, "marca"), LIMITE_TEXTO_CURTO),
            "categoria": _limpa_texto(celula(linha, "categoria"), LIMITE_TEXTO_CURTO),
        })

    if descartadas > 20:
        res.avisos.append(f"...e outras {descartadas - 20} linhas ficaram de fora pelo mesmo tipo de problema.")
    if duplicadas > 10:
        res.avisos.append(f"...e outras {duplicadas - 10} linhas repetidas foram ignoradas.")
    if ean_descartados > 10:
        res.avisos.append(f"...e outros {ean_descartados - 10} códigos de barras foram descartados.")
    if truncou_no_teto:
        # Recusa explícita, não corte silencioso: cortar sem avisar produz uma
        # cesta certa respondendo outra pergunta.
        res.avisos.append(
            f"⚠️ A planilha passa de {MAX_ITENS:,} itens. Gravei os {MAX_ITENS:,} primeiros e "
            "parei — divida a planilha se precisar do resto.".replace(",", ".")
        )

    if not res.itens:
        raise PlanilhaInvalida(
            "Nenhuma linha aproveitável: toda linha precisa de descrição e de um "
            "preço maior que zero."
        )

    res.resumo = {
        "arquivo": nome_arquivo,
        "linha_cabecalho": idx_cab + 1,
        "itens": len(res.itens),
        "linhas_descartadas": descartadas,
        "linhas_repetidas": duplicadas,
        "com_ean": sum(1 for i in res.itens if i["ean"]),
        "com_marca": sum(1 for i in res.itens if i["marca"]),
        "ean_descartados": ean_descartados,
        "truncada": truncou_no_teto,
    }
    logger.info("[PLANILHA] %s | %d itens | %d descartadas | %d repetidas",
                nome_arquivo, len(res.itens), descartadas, duplicadas)
    return res
